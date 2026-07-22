#!/usr/bin/env python3
"""
Script pour créer un fichier Excel de test pour les paiements à partir du template téléchargé
Utilise openpyxl si disponible
"""

import sys
import os
from datetime import datetime

def create_excel_from_template(template_path, output_path, eleves, frais):
    """Crée un fichier Excel à partir du template avec des données de test"""
    try:
        import openpyxl
        
        # Charger le template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Ajouter des données de test (ligne 2)
        # Ordre des colonnes: DatePaiement, Montant, Devise, ModePaiement, NomCompletEleve, LibelleFrais
        today = datetime.now()
        ws.cell(row=2, column=1, value=today)  # DatePaiement
        ws.cell(row=2, column=2, value=50000)  # Montant
        ws.cell(row=2, column=3, value="CDF")  # Devise
        ws.cell(row=2, column=4, value="Espèces")  # ModePaiement
        ws.cell(row=2, column=5, value=eleves[0] if eleves else "ELEVE TEST 1")  # NomCompletEleve
        ws.cell(row=2, column=6, value=frais[0] if frais else "Minerval")  # LibelleFrais
        
        # Ajouter une deuxième ligne de test
        ws.cell(row=3, column=1, value=today)  # DatePaiement
        ws.cell(row=3, column=2, value=30000)  # Montant
        ws.cell(row=3, column=3, value="CDF")  # Devise
        ws.cell(row=3, column=4, value="Mobile Money")  # ModePaiement
        ws.cell(row=3, column=5, value=eleves[1] if len(eleves) > 1 else "ELEVE TEST 2")  # NomCompletEleve
        ws.cell(row=3, column=6, value=frais[1] if len(frais) > 1 else "Frais examen")  # LibelleFrais
        
        wb.save(output_path)
        print(f"✅ Fichier Excel créé avec openpyxl : {output_path}")
        return True
        
    except ImportError:
        print("⚠️  openpyxl non disponible")
        return False
    except Exception as e:
        print(f"❌ Erreur avec openpyxl : {e}")
        return False

if __name__ == "__main__":
    template = sys.argv[1] if len(sys.argv) > 1 else "template_paiements_v2.xlsx"
    output = sys.argv[2] if len(sys.argv) > 2 else "test_paiements_v2.xlsx"
    
    eleves = sys.argv[3].split(",") if len(sys.argv) > 3 else []
    frais = sys.argv[4].split(",") if len(sys.argv) > 4 else []
    
    if not os.path.exists(template):
        print(f"❌ Template introuvable : {template}")
        sys.exit(1)
    
    if create_excel_from_template(template, output, eleves, frais):
        sys.exit(0)
    else:
        print("❌ Impossible de créer le fichier Excel")
        sys.exit(1)

