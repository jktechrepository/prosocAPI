#!/usr/bin/env python3
"""
Script Python pour créer un fichier Excel de test avec des données valides et invalides
Nécessite : pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import sys

def create_test_excel_file(filename="test_inscriptions.xlsx"):
    """Crée un fichier Excel de test avec des données valides et invalides"""
    
    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscriptions"
    
    # Définir les en-têtes
    headers = [
        "Type", "IdEcole", "IdClasse", "IdAnneeScolaire", "DateInscription", "StatutInscription",
        "NomEleve", "PostnomEleve", "PrenomEleve", "GenreEleve", "DateNaissanceEleve",
        "LieuNaissanceEleve", "NationaliteEleve", "ProvinceEleve", "VilleEleve", "CommuneEleve",
        "QuartierEleve", "AvenueEleve", "NumeroEleve", "CommentaireEleve", "PhotoEleveUrl", "MatriculeEleve",
        "NomCompletTuteur", "GenreTuteur", "EmailTuteur", "TelephoneTuteur",
        "NomCompletRepresentant", "TelephoneRepresentant", "PhotoTuteurUrl", "PieceIdentiteTuteur",
        "IdEleveExistant", "IdTuteurExistant"
    ]
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données de test
    test_data = [
        # Ligne 2 : Données valides
        {
            "Type": "Inscription",
            "IdEcole": 1,
            "IdClasse": 1,
            "IdAnneeScolaire": 1,
            "DateInscription": datetime.now(),
            "StatutInscription": "En attente",
            "NomEleve": "KABEYA",
            "PostnomEleve": "MULENGA",
            "PrenomEleve": "Jean",
            "GenreEleve": "M",
            "DateNaissanceEleve": date(2010, 5, 15),
            "LieuNaissanceEleve": "Kinshasa",
            "NationaliteEleve": "Congolaise",
            "ProvinceEleve": "Kinshasa",
            "VilleEleve": "Kinshasa",
            "CommuneEleve": "Gombe",
            "QuartierEleve": "Centre-ville",
            "AvenueEleve": "Avenue de la Paix",
            "NumeroEleve": "123",
            "CommentaireEleve": "Élève test valide",
            "PhotoEleveUrl": None,
            "MatriculeEleve": None,
            "NomCompletTuteur": "KABEYA MULENGA Pierre",
            "GenreTuteur": "M",
            "EmailTuteur": "pierre.kabeya@example.com",
            "TelephoneTuteur": "+243900123456",
            "NomCompletRepresentant": None,
            "TelephoneRepresentant": None,
            "PhotoTuteurUrl": None,
            "PieceIdentiteTuteur": None,
            "IdEleveExistant": None,
            "IdTuteurExistant": None
        },
        # Ligne 3 : Données valides (autre élève)
        {
            "Type": "Inscription",
            "IdEcole": 1,
            "IdClasse": 1,
            "IdAnneeScolaire": 1,
            "DateInscription": datetime.now(),
            "StatutInscription": "En attente",
            "NomEleve": "MUKAMBA",
            "PostnomEleve": "KASONGO",
            "PrenomEleve": "Marie",
            "GenreEleve": "F",
            "DateNaissanceEleve": date(2011, 8, 20),
            "LieuNaissanceEleve": "Kinshasa",
            "NationaliteEleve": "Congolaise",
            "ProvinceEleve": "Kinshasa",
            "VilleEleve": "Kinshasa",
            "CommuneEleve": "Lingwala",
            "QuartierEleve": "Quartier Test",
            "AvenueEleve": "Avenue Test",
            "NumeroEleve": "456",
            "CommentaireEleve": "Élève test valide 2",
            "PhotoEleveUrl": None,
            "MatriculeEleve": None,
            "NomCompletTuteur": "MUKAMBA KASONGO Paul",
            "GenreTuteur": "M",
            "EmailTuteur": "paul.mukamba@example.com",
            "TelephoneTuteur": "+243900123457",
            "NomCompletRepresentant": None,
            "TelephoneRepresentant": None,
            "PhotoTuteurUrl": None,
            "PieceIdentiteTuteur": None,
            "IdEleveExistant": None,
            "IdTuteurExistant": None
        },
        # Ligne 4 : Données invalides (nom manquant)
        {
            "Type": "Inscription",
            "IdEcole": 1,
            "IdClasse": 1,
            "IdAnneeScolaire": 1,
            "DateInscription": datetime.now(),
            "StatutInscription": "En attente",
            "NomEleve": "",  # ❌ INVALIDE : Nom manquant
            "PostnomEleve": "TEST",
            "PrenomEleve": "Erreur",
            "GenreEleve": "M",
            "DateNaissanceEleve": date(2012, 1, 1),
            "LieuNaissanceEleve": "Kinshasa",
            "NationaliteEleve": "Congolaise",
            "ProvinceEleve": None,
            "VilleEleve": None,
            "CommuneEleve": None,
            "QuartierEleve": None,
            "AvenueEleve": None,
            "NumeroEleve": None,
            "CommentaireEleve": "Test avec erreur",
            "PhotoEleveUrl": None,
            "MatriculeEleve": None,
            "NomCompletTuteur": "TEST Tuteur",
            "GenreTuteur": "M",
            "EmailTuteur": "tuteur@example.com",
            "TelephoneTuteur": "+243900123458",
            "NomCompletRepresentant": None,
            "TelephoneRepresentant": None,
            "PhotoTuteurUrl": None,
            "PieceIdentiteTuteur": None,
            "IdEleveExistant": None,
            "IdTuteurExistant": None
        },
        # Ligne 5 : Données invalides (genre invalide)
        {
            "Type": "Inscription",
            "IdEcole": 1,
            "IdClasse": 1,
            "IdAnneeScolaire": 1,
            "DateInscription": datetime.now(),
            "StatutInscription": "En attente",
            "NomEleve": "INVALIDE",
            "PostnomEleve": "TEST",
            "PrenomEleve": "Genre",
            "GenreEleve": "X",  # ❌ INVALIDE : Genre doit être M ou F
            "DateNaissanceEleve": date(2012, 2, 2),
            "LieuNaissanceEleve": "Kinshasa",
            "NationaliteEleve": "Congolaise",
            "ProvinceEleve": None,
            "VilleEleve": None,
            "CommuneEleve": None,
            "QuartierEleve": None,
            "AvenueEleve": None,
            "NumeroEleve": None,
            "CommentaireEleve": "Test avec genre invalide",
            "PhotoEleveUrl": None,
            "MatriculeEleve": None,
            "NomCompletTuteur": "INVALIDE TEST Tuteur",
            "GenreTuteur": "M",
            "EmailTuteur": "invalide@example.com",
            "TelephoneTuteur": "+243900123459",
            "NomCompletRepresentant": None,
            "TelephoneRepresentant": None,
            "PhotoTuteurUrl": None,
            "PieceIdentiteTuteur": None,
            "IdEleveExistant": None,
            "IdTuteurExistant": None
        },
        # Ligne 6 : Données invalides (email invalide)
        {
            "Type": "Inscription",
            "IdEcole": 1,
            "IdClasse": 1,
            "IdAnneeScolaire": 1,
            "DateInscription": datetime.now(),
            "StatutInscription": "En attente",
            "NomEleve": "EMAIL",
            "PostnomEleve": "INVALIDE",
            "PrenomEleve": "Test",
            "GenreEleve": "M",
            "DateNaissanceEleve": date(2012, 3, 3),
            "LieuNaissanceEleve": "Kinshasa",
            "NationaliteEleve": "Congolaise",
            "ProvinceEleve": None,
            "VilleEleve": None,
            "CommuneEleve": None,
            "QuartierEleve": None,
            "AvenueEleve": None,
            "NumeroEleve": None,
            "CommentaireEleve": "Test avec email invalide",
            "PhotoEleveUrl": None,
            "MatriculeEleve": None,
            "NomCompletTuteur": "EMAIL INVALIDE Tuteur",
            "GenreTuteur": "M",
            "EmailTuteur": "email-invalide",  # ❌ INVALIDE : Email mal formaté
            "TelephoneTuteur": "+243900123460",
            "NomCompletRepresentant": None,
            "TelephoneRepresentant": None,
            "PhotoTuteurUrl": None,
            "PieceIdentiteTuteur": None,
            "IdEleveExistant": None,
            "IdTuteurExistant": None
        },
        # Ligne 7 : Doublon (même élève que ligne 2)
        {
            "Type": "Inscription",
            "IdEcole": 1,
            "IdClasse": 1,
            "IdAnneeScolaire": 1,
            "DateInscription": datetime.now(),
            "StatutInscription": "En attente",
            "NomEleve": "KABEYA",  # ❌ DOUBLON : Même élève que ligne 2
            "PostnomEleve": "MULENGA",
            "PrenomEleve": "Jean",
            "GenreEleve": "M",
            "DateNaissanceEleve": date(2010, 5, 15),
            "LieuNaissanceEleve": "Kinshasa",
            "NationaliteEleve": "Congolaise",
            "ProvinceEleve": None,
            "VilleEleve": None,
            "CommuneEleve": None,
            "QuartierEleve": None,
            "AvenueEleve": None,
            "NumeroEleve": None,
            "CommentaireEleve": "Test doublon",
            "PhotoEleveUrl": None,
            "MatriculeEleve": None,
            "NomCompletTuteur": "KABEYA MULENGA Pierre",
            "GenreTuteur": "M",
            "EmailTuteur": "pierre.kabeya@example.com",
            "TelephoneTuteur": "+243900123456",
            "NomCompletRepresentant": None,
            "TelephoneRepresentant": None,
            "PhotoTuteurUrl": None,
            "PieceIdentiteTuteur": None,
            "IdEleveExistant": None,
            "IdTuteurExistant": None
        }
    ]
    
    # Écrire les données
    for row_idx, data in enumerate(test_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = data.get(header, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style pour les lignes avec erreurs
            if row_idx >= 4:  # Lignes 4-7 ont des erreurs
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ajuster la largeur des colonnes
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header), 15)
    
    # Ajouter une note
    note_row = len(test_data) + 3
    ws.cell(row=note_row, column=1, value="Note :")
    ws.cell(row=note_row, column=2, value="Lignes 2-3 : Données valides")
    ws.cell(row=note_row + 1, column=2, value="Ligne 4 : Nom manquant")
    ws.cell(row=note_row + 2, column=2, value="Ligne 5 : Genre invalide")
    ws.cell(row=note_row + 3, column=2, value="Ligne 6 : Email invalide")
    ws.cell(row=note_row + 4, column=2, value="Ligne 7 : Doublon (même élève que ligne 2)")
    
    # Sauvegarder le fichier
    wb.save(filename)
    print(f"✅ Fichier Excel de test créé : {filename}")
    print(f"   - {len(test_data)} lignes de données")
    print(f"   - 2 lignes valides (lignes 2-3)")
    print(f"   - 4 lignes avec erreurs (lignes 4-7)")
    print(f"   - 1 doublon (ligne 7)")
    
    return filename

if __name__ == "__main__":
    filename = sys.argv[1] if len(sys.argv) > 1 else "test_inscriptions.xlsx"
    try:
        create_test_excel_file(filename)
    except ImportError:
        print("❌ Erreur : openpyxl n'est pas installé")
        print("   Installez-le avec : pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Erreur : {e}")
        sys.exit(1)



