#!/usr/bin/env python3
"""
Script Python pour créer un fichier Excel de test avec des données valides et invalides pour les paiements
Nécessite : pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import sys

def create_test_excel_file(filename="test_paiements.xlsx"):
    """Crée un fichier Excel de test avec des données valides et invalides"""
    
    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paiements"
    
    # Définir les en-têtes
    headers = [
        "DatePaiement", "Montant", "Devise", "ModePaiement", "StatutPaiement",
        "ReferenceTransaction", "JustificatifUrl", "Commentaire",
        "IdEleve", "IdFrais", "IdUtilisateur", "Statut"
    ]
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données de test
    test_data = [
        # Ligne 2 : Données valides
        {
            "DatePaiement": datetime.now(),
            "Montant": 100.00,
            "Devise": "USD",
            "ModePaiement": "Cash",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-001",
            "JustificatifUrl": None,
            "Commentaire": "Paiement test valide",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 3 : Données valides (autre paiement)
        {
            "DatePaiement": datetime.now(),
            "Montant": 150.50,
            "Devise": "CDF",
            "ModePaiement": "Mobile Money",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-002",
            "JustificatifUrl": None,
            "Commentaire": "Paiement test valide 2",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 4 : Données invalides (montant manquant)
        {
            "DatePaiement": datetime.now(),
            "Montant": None,  # ❌ INVALIDE : Montant manquant
            "Devise": "USD",
            "ModePaiement": "Cash",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-003",
            "JustificatifUrl": None,
            "Commentaire": "Test avec montant manquant",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 5 : Données invalides (montant négatif)
        {
            "DatePaiement": datetime.now(),
            "Montant": -50.00,  # ❌ INVALIDE : Montant négatif
            "Devise": "USD",
            "ModePaiement": "Cash",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-004",
            "JustificatifUrl": None,
            "Commentaire": "Test avec montant négatif",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 6 : Données invalides (devise invalide)
        {
            "DatePaiement": datetime.now(),
            "Montant": 200.00,
            "Devise": "XYZ",  # ❌ INVALIDE : Devise non supportée
            "ModePaiement": "Cash",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-005",
            "JustificatifUrl": None,
            "Commentaire": "Test avec devise invalide",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 7 : Données invalides (mode de paiement invalide)
        {
            "DatePaiement": datetime.now(),
            "Montant": 75.00,
            "Devise": "USD",
            "ModePaiement": "Bitcoin",  # ❌ INVALIDE : Mode non supporté
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-006",
            "JustificatifUrl": None,
            "Commentaire": "Test avec mode de paiement invalide",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 8 : Données invalides (IdEleve manquant)
        {
            "DatePaiement": datetime.now(),
            "Montant": 300.00,
            "Devise": "USD",
            "ModePaiement": "Cash",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-007",
            "JustificatifUrl": None,
            "Commentaire": "Test avec IdEleve manquant",
            "IdEleve": None,  # ❌ INVALIDE : IdEleve manquant
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        },
        # Ligne 9 : Doublon (même paiement que ligne 2)
        {
            "DatePaiement": datetime.now(),  # ❌ DOUBLON : Même paiement que ligne 2
            "Montant": 100.00,
            "Devise": "USD",
            "ModePaiement": "Cash",
            "StatutPaiement": "Confirmé",
            "ReferenceTransaction": "REF-001",
            "JustificatifUrl": None,
            "Commentaire": "Test doublon",
            "IdEleve": 1,
            "IdFrais": 1,
            "IdUtilisateur": None,
            "Statut": True
        }
    ]
    
    # Écrire les données
    for row_idx, data in enumerate(test_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = data.get(header, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style pour les lignes avec erreurs
            if row_idx >= 4:  # Lignes 4-9 ont des erreurs
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ajuster la largeur des colonnes
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header), 15)
    
    # Ajouter une note
    note_row = len(test_data) + 3
    ws.cell(row=note_row, column=1, value="Note :")
    ws.cell(row=note_row, column=2, value="Lignes 2-3 : Données valides")
    ws.cell(row=note_row + 1, column=2, value="Ligne 4 : Montant manquant")
    ws.cell(row=note_row + 2, column=2, value="Ligne 5 : Montant négatif")
    ws.cell(row=note_row + 3, column=2, value="Ligne 6 : Devise invalide")
    ws.cell(row=note_row + 4, column=2, value="Ligne 7 : Mode de paiement invalide")
    ws.cell(row=note_row + 5, column=2, value="Ligne 8 : IdEleve manquant")
    ws.cell(row=note_row + 6, column=2, value="Ligne 9 : Doublon (même paiement que ligne 2)")
    
    # Sauvegarder le fichier
    wb.save(filename)
    print(f"✅ Fichier Excel de test créé : {filename}")
    print(f"   - {len(test_data)} lignes de données")
    print(f"   - 2 lignes valides (lignes 2-3)")
    print(f"   - 6 lignes avec erreurs (lignes 4-9)")
    print(f"   - 1 doublon (ligne 9)")
    
    return filename

if __name__ == "__main__":
    filename = sys.argv[1] if len(sys.argv) > 1 else "test_paiements.xlsx"
    try:
        create_test_excel_file(filename)
    except ImportError:
        print("❌ Erreur : openpyxl n'est pas installé")
        print("   Installez-le avec : pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Erreur : {e}")
        sys.exit(1)



