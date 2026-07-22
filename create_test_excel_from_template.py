#!/usr/bin/env python3
"""
Script pour créer un fichier Excel de test à partir du template téléchargé
Utilise openpyxl si disponible, sinon crée un fichier Excel simple
"""

import sys
import os
from datetime import datetime

def create_excel_simple(template_path, output_path, classes, annees):
    """Crée un fichier Excel simple en copiant le template et en ajoutant des données"""
    try:
        import openpyxl
        
        # Charger le template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Ajouter des données de test (ligne 2)
        today = datetime.now()
        ws.cell(row=2, column=1, value="Inscription")
        ws.cell(row=2, column=2, value=today)
        ws.cell(row=2, column=3, value="MUKENDI")
        ws.cell(row=2, column=4, value="KALALA")
        ws.cell(row=2, column=5, value="Jean")
        ws.cell(row=2, column=6, value="M")
        ws.cell(row=2, column=7, value=datetime(2010, 5, 15))
        ws.cell(row=2, column=8, value="Kinshasa")
        ws.cell(row=2, column=9, value="Congolaise")
        ws.cell(row=2, column=10, value="MUKENDI Pierre")
        ws.cell(row=2, column=11, value="M")
        ws.cell(row=2, column=12, value=classes[0] if classes else "1ère Primaire")
        ws.cell(row=2, column=13, value=annees[0] if annees else "2025-2026")
        
        # Ajouter une deuxième ligne de test
        ws.cell(row=3, column=1, value="Inscription")
        ws.cell(row=3, column=2, value=today)
        ws.cell(row=3, column=3, value="KALALA")
        ws.cell(row=3, column=4, value="MUKENDI")
        ws.cell(row=3, column=5, value="Marie")
        ws.cell(row=3, column=6, value="F")
        ws.cell(row=3, column=7, value=datetime(2011, 3, 20))
        ws.cell(row=3, column=8, value="Kinshasa")
        ws.cell(row=3, column=9, value="Congolaise")
        ws.cell(row=3, column=10, value="KALALA Marie")
        ws.cell(row=3, column=11, value="F")
        ws.cell(row=3, column=12, value=classes[1] if len(classes) > 1 else "2 eme Primaire")
        ws.cell(row=3, column=13, value=annees[1] if len(annees) > 1 else "2025-2026")
        
        wb.save(output_path)
        print(f"✅ Fichier Excel créé avec openpyxl : {output_path}")
        return True
        
    except ImportError:
        print("⚠️  openpyxl non disponible, création d'un fichier Excel basique...")
        return False
    except Exception as e:
        print(f"❌ Erreur avec openpyxl : {e}")
        return False

if __name__ == "__main__":
    template = sys.argv[1] if len(sys.argv) > 1 else "template_inscriptions_v2.xlsx"
    output = sys.argv[2] if len(sys.argv) > 2 else "test_inscriptions_v2.xlsx"
    
    classes = sys.argv[3].split(",") if len(sys.argv) > 3 else ["1ère Primaire", "2 eme Primaire"]
    annees = sys.argv[4].split(",") if len(sys.argv) > 4 else ["2025-2026", "2025-2026"]
    
    if not os.path.exists(template):
        print(f"❌ Template introuvable : {template}")
        sys.exit(1)
    
    if create_excel_simple(template, output, classes, annees):
        sys.exit(0)
    else:
        print("❌ Impossible de créer le fichier Excel")
        sys.exit(1)




